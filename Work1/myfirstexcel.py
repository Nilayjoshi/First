import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from google_api import *
from Google import Create_Service

exceldata = pd.read_excel("Client.xlsx")                                                                            
dfdata = pd.DataFrame(exceldata)

#print(dfdata)
df = dfdata.iloc[21]
#print(df)
df = dfdata.iloc[5][3]
#print(df2)

scope = ['https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive"]

credentials = ServiceAccountCredentials.from_json_keyfile_name('python-first.json', scope)
client = gspread.authorize(credentials)

sheets_service = create_sheets_service()
drive_service = create_drive_service()

spreadsheet_id = '1kM_sb3T_0ANmnuErCUxEoernRLUJbvDkURexpj36tX8'
sheet_to_excel(drive_service, spreadsheet_id)

sheet = client.open("FirstSheet").sheet1


worksheet_name = "Sheet1"
#cell_range_insert = "F 11"
li = ["A","B","C","D","E","F","G"]
value = (li)
a = 11
"""
update_values = []
for edit_cell in range(len(li)):
    value = li[edit_cell]
    update_values.append({'range' : worksheet_name + '!' + 'F' + str(a), 'values' : [[value]]})
     a+=1
"""
i = 0
update_values=[]
for j in range(65,74):
    value = dfdata.iloc[20][i]
    value = str(value)
    update_values.append({'range' : worksheet_name + '!' + str(chr(j)) + "22", 'values' : [[value]]})
    i+=1

batch_update_values_request_body = {
	'value_input_option': 'USER_ENTERED',
	'data': update_values,
}
request = sheets_service.spreadsheets().values().batchUpdate(spreadsheetId=spreadsheet_id, body=batch_update_values_request_body)
response = request.execute()














#'https://docs.google.com/spreadsheets/d/1kM_sb3T_0ANmnuErCUxEoernRLUJbvDkURexpj36tX8/edit#gid=0'
"""
import openpyxl 

worksheet = openpyxl.load_workbook("Client.xlsx")
sheet = worksheet.active
for value in sheet.iter_cols(values_only = True):
    ID=[8]
    if ID == "1ibHiHqvXpSjzisxAQAe7mi0rqyCCuq2bDevS3lV9pzU":
            print(dfdata.iloc[value])        
"""